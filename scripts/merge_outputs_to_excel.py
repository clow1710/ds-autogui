from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_DIR = PROJECT_ROOT / "tasks" / "realtask-20260513-2" / "outputs"

INVISIBLES = "".join(chr(cp) for cp in (0xFEFF, 0x200B, 0x200C, 0x200D, 0x2060, 0x00A0))

COLUMNS: tuple[tuple[str, str], ...] = (
    ("id", "id"),
    ("company_name", "company_name"),
    ("registered_address", "registered_address"),
    ("is_non_public", "is_non_public"),
    ("batch_id", "batch_id"),
    ("task_id", "task_id"),
    ("account_id", "account_id"),
    ("request_time", "request_time"),
    ("response_time", "response_time"),
    ("source_file", "source_file"),
)

COLUMN_WIDTHS: dict[str, int] = {
    "id": 10,
    "company_name": 38,
    "registered_address": 60,
    "is_non_public": 14,
    "batch_id": 28,
    "task_id": 28,
    "account_id": 10,
    "request_time": 22,
    "response_time": 22,
    "source_file": 36,
}


def _strip_invisibles(text: str) -> str:
    return text.strip().strip(INVISIBLES).strip()


def _parse_response(response: Any) -> dict | None:
    if not isinstance(response, str):
        return None
    s = _strip_invisibles(response)
    if not s or not s.startswith("{") or not s.endswith("}"):
        return None
    try:
        parsed = json.loads(s)
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, dict) else None


def _sort_key(row: dict) -> tuple[int, str, str]:
    cid_raw = str(row.get("id") or "")
    try:
        cid_num = int(cid_raw)
        return (0, "", f"{cid_num:020d}")
    except ValueError:
        return (1, cid_raw, "")


def collect_rows(input_dir: Path) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    warnings: list[str] = []
    json_paths = sorted(input_dir.glob("*.json"))
    if not json_paths:
        warnings.append(f"目录中没有 .json 文件：{input_dir}")
        return rows, warnings

    for path in json_paths:
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            warnings.append(f"跳过 {path.name}：外层 JSON 解析失败 ({exc})")
            continue

        parsed = _parse_response(data.get("response"))
        if parsed is None:
            warnings.append(f"跳过 {path.name}：response 字段不可解析为 JSON 对象")
            continue

        items = parsed.get("items")
        if not isinstance(items, list):
            warnings.append(f"跳过 {path.name}：response.items 不是数组")
            continue

        base = {
            "task_id": data.get("task_id"),
            "account_id": data.get("account_id"),
            "request_time": data.get("request_time"),
            "response_time": data.get("response_time"),
            "batch_id": parsed.get("batch_id"),
            "source_file": path.name,
        }

        for index, item in enumerate(items):
            if not isinstance(item, dict):
                warnings.append(f"{path.name} items[{index}] 不是对象，已跳过")
                continue
            row = {
                **base,
                "id": item.get("id"),
                "company_name": item.get("company_name"),
                "registered_address": item.get("registered_address"),
                "is_non_public": item.get("is_non_public"),
            }
            rows.append(row)

    rows.sort(key=_sort_key)
    return rows, warnings


def write_workbook(rows: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "merged"

    headers = [header for header, _ in COLUMNS]
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 18)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(key) for _, key in COLUMNS])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="将 DeepSeek 任务输出目录下的 JSON 合并为单个 Excel 表。"
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help=f"包含 *.json 响应的目录。默认：{DEFAULT_INPUT_DIR}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出 xlsx 路径。默认为 <input-dir>/../merged.xlsx",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_dir: Path = args.input_dir
    if not input_dir.is_dir():
        print(f"输入目录不存在：{input_dir}", file=sys.stderr)
        return 1

    output_path: Path = args.output or (input_dir.parent / "merged.xlsx")

    rows, warnings = collect_rows(input_dir)
    for msg in warnings:
        print(msg, file=sys.stderr)

    if not rows:
        print("没有可写入的记录。", file=sys.stderr)
        return 1

    write_workbook(rows, output_path)
    print(f"已写入 {len(rows)} 条记录到 {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
